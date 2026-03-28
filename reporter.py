"""
reporter.py — Excel Workpaper Generator
7 sheets: Cover · Ratios · Variance · Flagged · Going Concern · Benford's · Audit Notes
openpyxl loaded lazily inside generate_report to prevent startup crashes.
"""

import pandas as pd
from datetime import date

P = {
    "navy":    "0D1B2A", "navy2":  "1E3A5F", "slate":  "334155",
    "red_bg":  "FEE2E2", "yel_bg": "FEF9C3", "grn_bg": "DCFCE7",
    "grey":    "F8FAFC", "white":  "FFFFFF", "div":    "E2E8F0",
    "text":    "0F172A", "sub":    "475569", "light":  "94A3B8",
}


def generate_report(df_variance, ratios, agg, gc, bf,
                    output_path, client_name, period, standard, unit_label):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def bdr(c="E2E8F0"):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def flag_bg(f):
        s = str(f)
        if "🔴" in s: return P["red_bg"]
        if "🟡" in s: return P["yel_bg"]
        if "🟢" in s: return P["grn_bg"]
        return P["white"]

    def cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def rh(ws, row, h):
        ws.row_dimensions[row].height = h

    def hdr(ws, row, col, val, bg=None, fg="FFFFFF", sz=10, bold=True, span=1, wrap=False):
        bg = bg or P["navy"]
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=wrap)
        c.border = bdr()
        return c

    def cell(ws, row, col, val, bg=P["white"], bold=False, align="left", fmt=None, sz=10, wrap=False, fg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=sz, bold=bold, color=fg or P["text"])
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                indent=1 if align == "left" else 0, wrap_text=wrap)
        c.border = bdr()
        if fmt: c.number_format = fmt
        return c

    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════════════════════════════
    # SHEET 1 — COVER
    # ════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Cover")
    ws1.sheet_view.showGridLines = False
    N = 6

    ws1.merge_cells(f"A1:F1")
    t = ws1["A1"]
    t.value = "AUDIT ANALYTICAL PROCEDURES WORKPAPER"
    t.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    t.fill = fill(P["navy"]); t.border = bdr()
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws1, 1, 44)

    ws1.merge_cells("A2:F2")
    s = ws1["A2"]
    s.value = f"{standard}  ·  Analytical Procedures  ·  Going Concern  ·  Fraud (Benford's)"
    s.font = Font(name="Calibri", size=9, italic=True, color=P["light"])
    s.fill = fill(P["navy2"]); s.border = bdr()
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws1, 2, 16); rh(ws1, 3, 8)

    hdr(ws1, 4, 1, "ENGAGEMENT DETAILS", bg=P["navy2"], span=N, sz=9); rh(ws1, 4, 20)

    details = [
        ("Client", client_name), ("Period / Year End", period),
        ("Reporting Standard", standard), ("Amounts In", unit_label),
        ("Report Date", date.today().strftime("%d %B %Y")),
        ("Prepared By", "Audit AP Tool — Built by Purva Doshi  |  linkedin.com/in/purvadoshi26"),
    ]
    for i, (lbl, val) in enumerate(details, 5):
        rh(ws1, i, 18)
        lc = ws1.cell(row=i, column=1, value=lbl)
        lc.font = Font(name="Calibri", size=10, bold=True, color=P["text"])
        lc.fill = fill(P["grey"]); lc.border = bdr()
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=N)
        vc = ws1.cell(row=i, column=2, value=val)
        vc.font = Font(name="Calibri", size=10, color=P["text"])
        vc.fill = fill(P["white"]); vc.border = bdr()
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    rh(ws1, 12, 10)
    hdr(ws1, 13, 1, "RISK SUMMARY", bg=P["navy2"], span=N, sz=9); rh(ws1, 13, 20)

    gc_risk   = gc["overall_risk"]
    flag_high = df_variance["Flag"].str.contains("🔴", na=False).sum()
    flag_mod  = df_variance["Flag"].str.contains("🟡", na=False).sum()
    avail_r   = sum(1 for r in ratios if r.get("Available"))
    flag_r    = sum(1 for r in ratios if r.get("Available") and "🔴" in str(r.get("Flag", "")))
    total_rs  = df_variance["Risk Score"].sum() if "Risk Score" in df_variance.columns else 0

    summary_rows = [
        ("Overall Risk Level", gc_risk),
        ("Going Concern", f"{gc_risk} — Score {gc['score']}/10"),
        ("Variance — HIGH Risk Items", str(int(flag_high))),
        ("Variance — MODERATE Items", str(int(flag_mod))),
        ("Ratios Available / Flagged", f"{avail_r} calculated  ·  {flag_r} flagged HIGH"),
        ("Total Account Risk Score", str(int(total_rs))),
        ("Benford's Law", bf.get("risk_flag", "Not Analysed")),
    ]
    for i, (lbl, val) in enumerate(summary_rows, 14):
        rh(ws1, i, 18)
        lc = ws1.cell(row=i, column=1, value=lbl)
        lc.font = Font(name="Calibri", size=10, bold=True, color=P["text"])
        lc.fill = fill(P["grey"]); lc.border = bdr()
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=N)
        bg_v = flag_bg(val) if any(e in str(val) for e in ["🔴","🟡","🟢","CRITICAL","HIGH","MODERATE","LOW"]) else P["white"]
        if "CRITICAL" in str(val): bg_v = P["red_bg"]
        elif "HIGH" in str(val) and "Score" not in str(val): bg_v = P["red_bg"]
        elif "MODERATE" in str(val) and "Score" not in str(val): bg_v = P["yel_bg"]
        elif "LOW" in str(val) and "Score" not in str(val): bg_v = P["grn_bg"]
        vc = ws1.cell(row=i, column=2, value=val)
        vc.font = Font(name="Calibri", size=10, color=P["text"])
        vc.fill = fill(bg_v); vc.border = bdr()
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for c, w in [(1, 32), (2, 55), (3, 10), (4, 10), (5, 10), (6, 10)]:
        cw(ws1, c, w)

    # ════════════════════════════════════════════════════════
    # SHEET 2 — RATIO ANALYSIS
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Ratio Analysis")
    ws2.sheet_view.showGridLines = False
    h2 = ["Ratio / Metric", "Formula", f"CY ({unit_label})", f"PY ({unit_label})",
          "YoY (%)", "Flag", "Auditor Note", "Ref"]
    w2 = [30, 40, 13, 13, 10, 26, 55, 18]

    hdr(ws2, 1, 1, f"RATIO ANALYSIS — {standard}", span=8, sz=12); rh(ws2, 1, 28)
    rh(ws2, 2, 22)
    for col, (h, w) in enumerate(zip(h2, w2), 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = fill(P["slate"]); c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cw(ws2, col, w)

    sections = {"Profitability": "PROFITABILITY", "Liquidity": "LIQUIDITY",
                "Leverage": "LEVERAGE & SOLVENCY", "Efficiency": "EFFICIENCY"}
    last_sec = None
    data_r = 3
    for r in [x for x in ratios if x.get("Available")]:
        sec = r.get("Section", "")
        if sec != last_sec and sec in sections:
            ws2.merge_cells(start_row=data_r, start_column=1, end_row=data_r, end_column=8)
            sc = ws2.cell(row=data_r, column=1, value=f"  ── {sections[sec]}")
            sc.font = Font(name="Calibri", size=9, bold=True, color=P["sub"])
            sc.fill = fill(P["grey"]); sc.border = bdr()
            rh(ws2, data_r, 15); data_r += 1; last_sec = sec

        bg = flag_bg(r.get("Flag", ""))
        rh(ws2, data_r, 20)
        vals = [r.get("Ratio"), r.get("Formula"), r.get("CY"), r.get("PY"),
                r.get("YoY (%)"), r.get("Flag"), r.get("Note"), r.get("Ref")]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=data_r, column=col, value=val)
            c.font = Font(name="Calibri", size=9, bold=(col == 1), color=P["text"])
            c.fill = fill(bg); c.border = bdr()
            c.alignment = Alignment(horizontal="right" if col in [3,4,5] else "left",
                                    vertical="center", wrap_text=True, indent=0 if col in [3,4,5] else 1)
            if col in [3, 4] and isinstance(val, float): c.number_format = "0.00"
            if col == 5 and isinstance(val, float): c.number_format = "+0.0;-0.0"
        data_r += 1

    ws2.freeze_panes = ws2.cell(row=3, column=1)

    # ════════════════════════════════════════════════════════
    # SHEET 3 — ALL ACCOUNTS
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Variance — All Accounts")
    ws3.sheet_view.showGridLines = False
    h3 = ["Account Name", "Category", f"CY ({unit_label})", f"PY ({unit_label})",
          f"Change ({unit_label})", "Change (%)", "Material?", "Flag", "Risk Score", "Audit Commentary"]
    w3 = [38, 22, 14, 14, 14, 11, 10, 26, 10, 60]

    hdr(ws3, 1, 1, f"VARIANCE ANALYSIS — All Accounts  |  {standard}", span=10, sz=11); rh(ws3, 1, 26)
    rh(ws3, 2, 20)
    for col, (h, w) in enumerate(zip(h3, w3), 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = fill(P["slate"]); c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cw(ws3, col, w)

    for i, (_, row) in enumerate(df_variance.iterrows(), 3):
        flag = str(row.get("Flag", ""))
        bg   = flag_bg(flag)
        if bg == P["white"]: bg = P["grey"] if i % 2 == 0 else P["white"]
        rh(ws3, i, 38)
        vals = [row["Account Name"], row["Category"], row["CY Amount"], row["PY Amount"],
                row["Change"], row["Change (%)"], row.get("Material?", ""),
                flag, row.get("Risk Score", 0), row.get("Audit Commentary", "")]
        for col, val in enumerate(vals, 1):
            disp = "New" if col == 6 and val == 999.0 else val
            c = ws3.cell(row=i, column=col, value=disp)
            c.font = Font(name="Calibri", size=9, color=P["text"])
            c.fill = fill(bg); c.border = bdr()
            c.alignment = Alignment(
                horizontal="right" if col in [3,4,5,6,9] else "left",
                vertical="top", wrap_text=True,
                indent=1 if col in [1,2,7,8,10] else 0)
            if col in [3,4,5] and isinstance(val, (int,float)): c.number_format = "#,##0.00"
            if col == 6 and isinstance(val, float) and val != 999.0: c.number_format = "+0.0;-0.0"

    ws3.freeze_panes = ws3.cell(row=3, column=1)

    # ════════════════════════════════════════════════════════
    # SHEET 4 — FLAGGED ITEMS
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Flagged Items")
    ws4.sheet_view.showGridLines = False
    flagged = df_variance[df_variance["Flag"].str.contains("🔴|🟡", na=False)]

    hdr(ws4, 1, 1, f"FLAGGED ITEMS — Requires Auditor Attention  |  {standard}", span=10, sz=11); rh(ws4, 1, 26)
    rh(ws4, 2, 20)
    for col, (h, w) in enumerate(zip(h3, w3), 1):
        c = ws4.cell(row=2, column=col, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = fill(P["slate"]); c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cw(ws4, col, w)

    if flagged.empty:
        ws4.merge_cells("A3:J3")
        nc = ws4["A3"]
        nc.value = "✅ No flagged items — all accounts within materiality threshold."
        nc.font = Font(name="Calibri", size=10, color="14532D")
        nc.fill = fill(P["grn_bg"]); nc.border = bdr()
        rh(ws4, 3, 22)
    else:
        for i, (_, row) in enumerate(flagged.iterrows(), 3):
            flag = str(row.get("Flag", ""))
            bg   = flag_bg(flag)
            rh(ws4, i, 38)
            vals = [row["Account Name"], row["Category"], row["CY Amount"], row["PY Amount"],
                    row["Change"], row["Change (%)"], row.get("Material?", ""),
                    flag, row.get("Risk Score", 0), row.get("Audit Commentary", "")]
            for col, val in enumerate(vals, 1):
                disp = "New" if col == 6 and val == 999.0 else val
                c = ws4.cell(row=i, column=col, value=disp)
                c.font = Font(name="Calibri", size=9, color=P["text"])
                c.fill = fill(bg); c.border = bdr()
                c.alignment = Alignment(
                    horizontal="right" if col in [3,4,5,6,9] else "left",
                    vertical="top", wrap_text=True,
                    indent=1 if col in [1,2,7,8,10] else 0)
                if col in [3,4,5] and isinstance(val,(int,float)): c.number_format = "#,##0.00"
                if col == 6 and isinstance(val,float) and val != 999.0: c.number_format = "+0.0;-0.0"

    ws4.freeze_panes = ws4.cell(row=3, column=1)

    # ════════════════════════════════════════════════════════
    # SHEET 5 — GOING CONCERN
    # ════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Going Concern")
    ws5.sheet_view.showGridLines = False
    gc_col = {"CRITICAL": "991B1B", "HIGH": "C2410C", "MODERATE": "92400E", "LOW": "14532D"}
    gc_bg  = {"CRITICAL": P["red_bg"], "HIGH": "FEF3C7", "MODERATE": P["yel_bg"], "LOW": P["grn_bg"]}
    risk   = gc["overall_risk"]

    hdr(ws5, 1, 1, f"GOING CONCERN ASSESSMENT — {standard}", span=5, sz=12); rh(ws5, 1, 28)
    rh(ws5, 2, 8); rh(ws5, 3, 32)
    ws5.merge_cells("A3:E3")
    vc = ws5["A3"]
    vc.value = f"OVERALL: {risk}  ·  Risk Score: {gc['score']}/10"
    vc.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    vc.fill = fill(gc_col.get(risk, P["navy"])); vc.border = bdr()
    vc.alignment = Alignment(horizontal="center", vertical="center")

    rh(ws5, 4, 56)
    ws5.merge_cells("A4:E4")
    cc = ws5["A4"]
    cc.value = gc["conclusion"]
    cc.font = Font(name="Calibri", size=10, color=P["text"])
    cc.fill = fill(gc_bg.get(risk, P["grey"])); cc.border = bdr()
    cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    rh(ws5, 5, 10)
    hdr(ws5, 6, 1, "DETAILED INDICATORS", bg=P["slate"], span=5, sz=9); rh(ws5, 6, 20)
    for col, h in enumerate(["Indicator", "Status", "Finding", "Detail", "Reference"], 1):
        c = ws5.cell(row=7, column=col, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = fill(P["slate"]); c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws5, 7, 20)
    for c, w in [(1,38),(2,22),(3,44),(4,36),(5,28)]:
        cw(ws5, c, w)

    for i, ind in enumerate(gc["indicators"], 8):
        flag = str(ind.get("Status", ""))
        bg   = flag_bg(flag)
        rh(ws5, i, 34)
        for col, val in enumerate([ind["Indicator"], ind["Status"], ind["Finding"],
                                   ind.get("Detail",""), ind["Reference"]], 1):
            c = ws5.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=9, bold=(col==1), color=P["text"])
            c.fill = fill(bg); c.border = bdr()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

    # ════════════════════════════════════════════════════════
    # SHEET 6 — BENFORD'S LAW
    # ════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Benford's Law")
    ws6.sheet_view.showGridLines = False
    hdr(ws6, 1, 1, "BENFORD'S LAW — ISA/SA 240 Fraud Risk Indicator", span=7, sz=12); rh(ws6, 1, 28)

    if not bf.get("sufficient"):
        ws6.merge_cells("A2:G2")
        nc = ws6["A2"]
        nc.value = f"NOT ANALYSED — {bf.get('message', 'No transaction data uploaded.')}"
        nc.font = Font(name="Calibri", size=10, color=P["text"])
        nc.fill = fill(P["yel_bg"]); nc.border = bdr()
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        rh(ws6, 2, 40)
    else:
        flag = bf.get("risk_flag", "")
        ws6.merge_cells("A2:G2")
        rc = ws6["A2"]
        rc.value = f"{flag}  ·  {bf['n']:,} amounts  ·  χ² = {bf['chi_square']:.2f}"
        rc.font = Font(name="Calibri", size=11, bold=True, color=P["text"])
        rc.fill = fill(flag_bg(flag)); rc.border = bdr()
        rc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws6, 2, 24)

        ws6.merge_cells("A3:G3")
        ic = ws6["A3"]
        ic.value = bf.get("interpretation", "")
        ic.font = Font(name="Calibri", size=10, color=P["text"])
        ic.fill = fill(P["grey"]); ic.border = bdr()
        ic.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        rh(ws6, 3, 52)

        rh(ws6, 4, 10)
        hdr(ws6, 5, 1, "FIRST-DIGIT ANALYSIS", bg=P["slate"], span=7, sz=9); rh(ws6, 5, 20)
        bf_h = ["Digit","Count","Observed (%)","Expected (%)","Deviation (pp)","Deviation (%)","Assessment"]
        bf_w = [10,12,14,22,15,14,22]
        for col, (h, w) in enumerate(zip(bf_h, bf_w), 1):
            c = ws6.cell(row=6, column=col, value=h)
            c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
            c.fill = fill(P["slate"]); c.border = bdr()
            c.alignment = Alignment(horizontal="center", vertical="center")
            cw(ws6, col, w)
        rh(ws6, 6, 20)

        for i, dr in enumerate(bf["digit_data"], 7):
            dev = abs(dr["Deviation (pp)"])
            df_bg = P["red_bg"] if dev>5 else P["yel_bg"] if dev>2 else P["grn_bg"]
            df_fl = "🔴 Investigate" if dev>5 else "🟡 Review" if dev>2 else "🟢 OK"
            rh(ws6, i, 18)
            for col, val in enumerate([dr["Digit"],dr["Observed Count"],dr["Observed (%)"],
                                       dr["Expected (%)"],dr["Deviation (pp)"],dr["Deviation (%)"],df_fl],1):
                c = ws6.cell(row=i, column=col, value=val)
                c.font = Font(name="Calibri", size=9, color=P["text"])
                c.fill = fill(df_bg); c.border = bdr()
                c.alignment = Alignment(horizontal="center", vertical="center")
                if col in [3,4]: c.number_format = "0.00"
                if col == 5: c.number_format = "+0.00;-0.00"

        hdr(ws6, 17, 1,
            f"χ² critical values (df=8): p<0.10 → 13.36  |  p<0.05 → 15.51  |  p<0.01 → 20.09  |  "
            f"Your statistic: {bf['chi_square']:.2f}",
            bg=P["slate"], span=7, sz=9)
        rh(ws6, 17, 20)

    # ════════════════════════════════════════════════════════
    # SHEET 7 — AUDIT NOTES (standards, no how-to content)
    # ════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Audit Notes")
    ws7.sheet_view.showGridLines = False
    hdr(ws7, 1, 1, f"AUDIT NOTES — Standards & Procedure References  |  {standard}", span=3, sz=12)
    rh(ws7, 1, 28)

    from engine import STD_REFS
    refs = STD_REFS.get(standard, STD_REFS["IFRS"])

    sections7 = [
        ("STANDARDS APPLIED", [
            ("Analytical Procedures", refs["ap"]),
            ("Going Concern", refs["gc"]),
            ("Fraud Risk (Benford's)", refs["fraud"]),
            ("Risk Assessment", refs["risk"]),
            ("Revenue Recognition", refs["revenue"]),
            ("Impairment", refs["impairment"]),
            ("Leases", refs["leases"]),
        ]),
        ("METHODOLOGY", [
            ("Variance — HIGH", "Change ≥ 1.5× threshold % AND above absolute threshold"),
            ("Variance — MODERATE", "Change ≥ threshold % AND above absolute threshold"),
            ("Materiality threshold", "Auto-scaled: ₹ Crores → 2.0  |  ₹ Lakhs → 5.0"),
            ("Going Concern Score", "0 = LOW  |  1–2 = MODERATE  |  3–5 = HIGH  |  6+ = CRITICAL"),
            ("Benford's Law", "χ² test df=8  |  p<0.05 → notable  |  p<0.01 → significant deviation"),
            ("Risk Score per account", "HIGH flag = +3  |  MODERATE flag = +1"),
        ]),
        ("WORKPAPER CONTENTS", [
            ("Sheet 1 — Cover", "Engagement details and risk summary"),
            ("Sheet 2 — Ratio Analysis", "15 financial ratios with risk flags and auditor notes"),
            ("Sheet 3 — Variance", "All accounts — YoY movement, materiality, audit commentary"),
            ("Sheet 4 — Flagged Items", "HIGH and MODERATE risk accounts only"),
            ("Sheet 5 — Going Concern", f"6 indicators assessed under {refs['gc']}"),
            ("Sheet 6 — Benford's Law", f"First-digit analysis under {refs['fraud']}"),
            ("Sheet 7 — Audit Notes", "This sheet — standard references and methodology"),
        ]),
        ("WORKPAPER DETAILS", [
            ("Client", client_name),
            ("Period", period),
            ("Standard", standard),
            ("Amounts", unit_label),
            ("Prepared", date.today().strftime("%d %B %Y")),
            ("Tool", "Audit AP Tool — Built by Purva Doshi"),
            ("LinkedIn", "https://www.linkedin.com/in/purvadoshi26/"),
        ]),
    ]

    row = 2
    for sec_title, items in sections7:
        rh(ws7, row, 8)
        row += 1
        hdr(ws7, row, 1, sec_title, bg=P["slate"], span=3, sz=9); rh(ws7, row, 20); row += 1
        for lbl, val in items:
            rh(ws7, row, 17)
            lc = ws7.cell(row=row, column=1, value=lbl)
            lc.font = Font(name="Calibri", size=10, bold=True, color=P["text"])
            lc.fill = fill(P["grey"]); lc.border = bdr()
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            vc = ws7.cell(row=row, column=2, value=val)
            vc.font = Font(name="Calibri", size=10, color=P["text"])
            vc.fill = fill(P["white"]); vc.border = bdr()
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row += 1

    for c, w in [(1, 36), (2, 55), (3, 20)]:
        cw(ws7, c, w)

    wb.save(output_path)
    return output_path
